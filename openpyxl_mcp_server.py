from datetime import datetime, timedelta, time
from pathlib import Path
import sys
from typing import Any

from mcp.server.fastmcp import FastMCP

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.datetime import from_excel


mcp = FastMCP("openpyxl_mcp_server")

OPENPYXL_TYPE_TO_STRING = {
    "n": "numeric",
    "s": "string",
    "f": "formula",
    "b": "boolean",
    "e": "error",
}

FILEPATH_DOCSTRING = 'The path to the Excel workbook. For example, "~/Downloads/test.xlsx" or "C:\\myfolder\\myfile.xlsx". If only a filename is provided, the file will be searched for in the Desktop and Downloads folders.'


@mcp.tool()
async def get_cell_details(filepath: str, sheet_name: str, cell_name: str) -> str:
    f"""Get value, data type, style, comments, formulas, hyperlinks, and other details for a single cell in a workbook.

    Args:
        file_path: {FILEPATH_DOCSTRING}
        sheet_name: The name of the sheet to get the value from.
        cell name: The name of the cell to get the value from. For example, "A1", "B2", "R5987".
    """
    filepath_clean = resolve_path_and_assert_file_exists(filepath)
    wb = load_workbook(filename=filepath_clean)
    sheet = get_sheet_and_assert_it_exists(wb, sheet_name)
    cell = sheet[cell_name]

    cellinfo = [
        f"Cell: {cell.coordinate}",
        f"Value: {cell.value}",
        f"Data type: {OPENPYXL_TYPE_TO_STRING.get(cell.data_type, cell.data_type)}",
        f"Style: {cell.style}",
        f"Number format: {cell.number_format}",
    ]

    # date values
    if cell.is_date:
        date_value: datetime | timedelta | time | None = from_excel(cell.value)
        if isinstance(date_value, (datetime, time)):
            cellinfo.append(f"Value as Date: {date_value.isoformat()}")
        elif isinstance(date_value, timedelta):
            cellinfo.append(
                f"Value as Time Interval: {date_value.total_seconds()} seconds"
            )

    # formulas
    if cell.data_type == "f":
        cellinfo.append(f"Formula: {cell.value}")

    # hyperlinks
    if cell.hyperlink:
        cellinfo.append(f"Hyperlink Text: {cell.hyperlink}")
        if hasattr(cell.hyperlink, "target"):
            cellinfo.append(f"Hyperlink Target: {cell.hyperlink.target}")
        if hasattr(cell.hyperlink, "tooltip"):
            cellinfo.append(f"Hyperlink Tooltip: {cell.hyperlink.tooltip}")

    # comment
    if cell.comment:
        cellinfo.append(f"Comment: {cell.comment.text}")
        cellinfo.append(f"Comment Author: {cell.comment.author}")

    # font
    try:
        font = cell.font
        cellinfo.append(f"Font Name: {font.name}")
        cellinfo.append(f"Font Size: {font.size}")
        cellinfo.append(f"Bold: {font.bold}")
        cellinfo.append(f"Italic: {font.italic}")
        cellinfo.append(f"Underline: {font.underline}")
        if font.color:
            if hasattr(font.color, "rgb") and font.color.rgb:
                cellinfo.append(f"Font Color (RGB): {font.color.rgb}")
            elif hasattr(font.color, "theme") and font.color.theme is not None:
                cellinfo.append(f"Font Color (Theme): {font.color.theme}")
            else:
                cellinfo.append(f"Font Color: {font.color}")
        else:
            cellinfo.append("Font Color: Default")
    except Exception:
        pass

    try:
        fill = cell.fill
        if hasattr(fill, "patternType") and fill.patternType:
            cellinfo.append(f"Fill Pattern Type: {fill.patternType}")

            if hasattr(fill, "fgColor") and fill.fgColor:
                if hasattr(fill.fgColor, "rgb") and fill.fgColor.rgb:
                    cellinfo.append(f"Fill Foreground Color (RGB): {fill.fgColor.rgb}")
                elif hasattr(fill.fgColor, "theme") and fill.fgColor.theme is not None:
                    cellinfo.append(
                        f"Fill Foreground Color (Theme): {fill.fgColor.theme}"
                    )
                else:
                    cellinfo.append(f"Fill Foreground Color: {fill.fgColor}")

            if hasattr(fill, "bgColor") and fill.bgColor:
                if hasattr(fill.bgColor, "rgb") and fill.bgColor.rgb:
                    cellinfo.append(f"Fill Background Color (RGB): {fill.bgColor.rgb}")
                elif hasattr(fill.bgColor, "theme") and fill.bgColor.theme is not None:
                    cellinfo.append(
                        f"Fill Background Color (Theme): {fill.bgColor.theme}"
                    )
                else:
                    cellinfo.append(f"Fill Background Color: {fill.bgColor}")
        else:
            cellinfo.append("Fill: No fill pattern")
    except Exception:
        pass

    try:
        alignment = cell.alignment
        cellinfo.append(f"Horizontal Alignment: {alignment.horizontal}")
        cellinfo.append(f"Vertical Alignment: {alignment.vertical}")
        cellinfo.append(f"Text Rotation: {alignment.textRotation}")
        cellinfo.append(f"Wrap Text: {alignment.wrapText}")
        cellinfo.append(f"Indent: {alignment.indent}")
        cellinfo.append(f"Shrink to Fit: {alignment.shrinkToFit}")
    except Exception:
        pass

    # border
    if cell.border:
        border_sides = {
            "left": cell.border.left,
            "right": cell.border.right,
            "top": cell.border.top,
            "bottom": cell.border.bottom,
            "diagonal": cell.border.diagonal,
        }

        for side_name, side in border_sides.items():
            b_name = f"Border {side_name.capitalize()}"
            if side and side.style:
                cellinfo.append(f"{b_name} Style: {side.style}")
                if side.color:
                    if hasattr(side.color, "rgb") and side.color.rgb:
                        cellinfo.append(f"{b_name} Color (RGB): {side.color.rgb}")
                    elif hasattr(side.color, "theme") and side.color.theme is not None:
                        cellinfo.append(f"{b_name} Color (Theme): {side.color.theme}")
                    else:
                        cellinfo.append(f"{b_name} Color: {side.color}")

    # protection
    if cell.protection:
        cellinfo.append(f"Is Cell Locked: {cell.protection.locked}")
        cellinfo.append(f"Is Cell Hidden: {cell.protection.hidden}")

    # conditional formatting
    cf_rules = []
    for rule in sheet.conditional_formatting:
        if f"P{cell.row}" in rule.cells.ranges:
            cf_rules.append(rule)

    if cf_rules:
        for i, rule in enumerate(cf_rules):
            cellinfo.append(f"Conditional Formatting Rule {i+1}:")
            for subrule in rule.rules:
                cellinfo.append(f"  Type: {type(subrule).__name__}")
                if hasattr(subrule, "formula"):
                    cellinfo.append(f"  Formula: {subrule.formula}")
                if hasattr(subrule, "operator"):
                    cellinfo.append(f"  Operator: {subrule.operator}")
                if hasattr(subrule, "dxf") and subrule.dxf:
                    cellinfo.append("  Differential Style:")
                    if hasattr(subrule.dxf, "font") and subrule.dxf.font:
                        cellinfo.append(f"    Font: {subrule.dxf.font}")
                    if hasattr(subrule.dxf, "fill") and subrule.dxf.fill:
                        cellinfo.append(f"    Fill: {subrule.dxf.fill}")
                    if hasattr(subrule.dxf, "border") and subrule.dxf.border:
                        cellinfo.append(f"    Border: {subrule.dxf.border}")

    # merged cells
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            cellinfo.append(f"Cell is part of merged range: {merged_range}")
            cellinfo.append(
                f"Merge starts at: {merged_range.min_row}, {merged_range.min_col}"
            )
            cellinfo.append(
                f"Merge ends at: {merged_range.max_row}, {merged_range.max_col}"
            )
            break

    return "\n".join(cellinfo)


@mcp.tool()
async def get_cell_value(filepath: str, sheet_name: str, cell_name: str) -> str:
    f"""Get the raw value of a single cell in a workbook.

    Args:
        file_path: {FILEPATH_DOCSTRING}
        sheet_name: The name of the sheet to get the value from.
        cell name: The name of the cell to get the value from. For example, "A1", "B2", "R5987".
    """
    filepath_clean = resolve_path_and_assert_file_exists(filepath)
    wb = load_workbook(filename=filepath_clean)
    sheet = get_sheet_and_assert_it_exists(wb, sheet_name)
    cell = sheet[cell_name]
    return str({cell.value})


@mcp.tool()
async def get_values_of_cell_range(
    filepath: str, sheet_name: str, top_left_cell: str, bottom_right_cell: str
) -> str:
    f"""Get the value, data type, style, and any comments, of a continuous range of cells in an Excel workbook.

    Args:
        file_path: {FILEPATH_DOCSTRING}
        sheet_name: The name of the sheet to get the value from.
        top_left_cell: The top left cell of the range. For example, "A1".
        bottom_right_cell: The bottom right cell of the range. For example, "RC976".
    """
    filepath_clean = resolve_path_and_assert_file_exists(filepath)
    wb = load_workbook(filename=filepath_clean)
    sheet = get_sheet_and_assert_it_exists(wb, sheet_name)
    range_str = ":".join(sorted([top_left_cell, bottom_right_cell]))
    cell_range = sheet[range_str]
    result = []
    for row in cell_range:
        for cell in row:
            result.append(f"{cell.coordinate}: {cell.value}")

    return "\n".join(result)


@mcp.tool()
async def get_content_of_cell_list(
    filepath: str, sheet_name: str, cell_name_list: list[str]
) -> str:
    f"""Get the raw values of a list of specific named cells in an Excel workbook.

    Args:
        file_path: {FILEPATH_DOCSTRING}
        sheet_name: The name of the sheet to get the value from.
        cell_name_list: A list of cell names. For example, ["A1", "B2", "C3"].
    """
    filepath_clean = resolve_path_and_assert_file_exists(filepath)
    wb = load_workbook(filename=filepath_clean)
    sheet = get_sheet_and_assert_it_exists(wb, sheet_name)
    result = []
    for cell_name in cell_name_list:
        cell = sheet[cell_name]
        result.append(f"{cell.coordinate}: {cell.value}")
    return "\n".join(result)


@mcp.tool()
async def search_in_cell_range(
    filepath: str,
    sheet_name: str,
    top_left_cell: str,
    bottom_right_cell: str,
    search_string: str,
    exact_match: bool = False,
) -> str:
    f"""Search for a string in a continuous range of cells in an Excel workbook.

    Args:
        file_path: {FILEPATH_DOCSTRING}
        sheet_name: The name of the sheet to get the value from.
        top_left_cell: The top left cell of the range. For example, "A1".
        bottom_right_cell: The bottom right cell of the range. For example, "RC976".
        search_string: The string to search for.
        exact_match: True if the entire cell value must match, False if it can be a substring. Defaults to False.
    """
    filepath_clean = resolve_path_and_assert_file_exists(filepath)
    wb = load_workbook(filename=filepath_clean)
    sheet = get_sheet_and_assert_it_exists(wb, sheet_name)
    range_str = ":".join(sorted([top_left_cell, bottom_right_cell]))
    cell_range = sheet[range_str]
    result = []
    for row in cell_range:
        for cell in row:
            if exact_match:
                if search_string == str(cell.value):
                    result.append(f"{cell.coordinate}: {cell.value}")
            else:
                if search_string in str(cell.value):
                    result.append(f"{cell.coordinate}: {cell.value}")

    return "\n".join(result)


@mcp.tool()
async def get_list_of_sheets(filepath: str) -> str:
    f"""Get a list of sheets in an Excel workbook. Each line contains a sheet's name and dimensions.

    Args:
        file_path: {FILEPATH_DOCSTRING}
    """
    filepath_clean = resolve_path_and_assert_file_exists(filepath)
    wb = load_workbook(filename=filepath_clean)
    result = []
    for sheet in wb.worksheets:
        result.append(f"Name: {sheet.title}, Dimensions: {sheet.dimensions}")
    return "\n".join(result)


def resolve_path_and_assert_file_exists(filepath: str) -> Path:
    expanded_path = Path(filepath).expanduser()
    if expanded_path.exists():
        return expanded_path
    # If filepath is just a filename and this is Winodws or MacOS, try the default locations of the Desktop and
    # Downloads directories
    is_windows = sys.platform == "win32" and "\\" not in filepath
    is_macos = sys.platform == "darwin" and "/" not in filepath
    if is_windows or is_macos:
        path_in_desktop = Path.home() / "Desktop" / filepath
        if path_in_desktop.exists():
            return path_in_desktop
        path_in_downloads = Path.home() / "Downloads" / filepath
        if path_in_downloads.exists():
            return path_in_downloads
    raise ValueError(f"File '{filepath}' does not exist")


def get_sheet_and_assert_it_exists(wb: Workbook, sheet_name: str) -> Worksheet:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} does not exist")
    return wb[sheet_name]


def render_cell_value(val: Any) -> str:
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(val, float):
        return f"{val:.2f}"
    else:
        return str(val)


if __name__ == "__main__":
    try:
        mcp.run(transport="stdio")
    except KeyboardInterrupt:
        pass
    except Exception as e:
        print(e, file=sys.stderr)
        raise
